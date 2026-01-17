import openpyxl

# Read main data file
wb = openpyxl.load_workbook(r'd:\Gadium\doc\excel pruebas\Satifacción Personal\Satisfacción Operarios.xlsx')
ws = wb.active

print("=== SATISFACCIÓN OPERARIOS (Main Data) ===")
print(f"Total columns: {ws.max_column}")
print(f"Total rows: {ws.max_row}")
print("\nHeaders:")
headers = [cell.value for cell in ws[1]]
for i, h in enumerate(headers, 1):
    print(f"{i}. {h}")

print("\nFirst data row:")
first_row = [cell.value for cell in ws[2]]
for i, val in enumerate(first_row, 1):
    print(f"{i}. {val}")

# Read results file
wb2 = openpyxl.load_workbook(r'd:\Gadium\doc\excel pruebas\Satifacción Personal\Satisfacción Operarios - resultados.xlsx')
ws2 = wb2.active

print("\n\n=== SATISFACCIÓN OPERARIOS - RESULTADOS ===")
print(f"Total rows: {ws2.max_row}")
for i in range(1, min(6, ws2.max_row + 1)):
    row_data = [cell.value for cell in ws2[i]]
    print(f"Row {i}: {row_data}")
